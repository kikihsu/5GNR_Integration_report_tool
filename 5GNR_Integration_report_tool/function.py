"""
KPI Log File Processing Tool - Function Library
===============================================
Core functions for reading, filtering, and processing log files.

Author: Kiki Hsu
Date: 2025-01-15
"""

import os
import datetime
import subprocess
from copy import copy

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from config import KPI_RULES, STYLE


# ============================================================================
# SHARED STYLE HELPERS
# ============================================================================

def _make_border(color="000000", style="thin"):
    side = Side(border_style=style, color=color)
    return Border(left=side, right=side, top=side, bottom=side)

def _make_fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def _make_font(hex_color="000000", bold=False, size=11):
    return Font(color=hex_color, bold=bold, size=size)

def _kpi_passes(rule, value):
    """
    Returns True if `value` satisfies the KPI rule's threshold.
    rule = (col, display, operator, threshold)
    """
    _, _, op, threshold = rule
    if op == "<":
        return value < threshold
    if op == ">":
        return value > threshold
    if op == "<=":
        return value <= threshold
    if op == ">=":
        return value >= threshold
    raise ValueError(f"Unknown operator: {op}")


# ============================================================================
# FILE READING
# ============================================================================

def read_file(file_path, filename):
    """
    Reads a single log file with automatic encoding detection.
    Returns a list of dicts (one per data row), or [] on failure.
    """
    print(f"  -> Reading: {filename}")
    data = []

    try:
        lines = _detect_and_read(file_path)
        data = _parse_log_lines(lines)

        if data:
            print(f"    Extracted: {len(data)} rows")
        else:
            print("    Warning: No data rows found")

    except FileNotFoundError:
        print(f"    ERROR: File not found at {file_path}")
    except Exception as exc:
        import traceback
        print(f"    ERROR: {exc}")
        traceback.print_exc()

    return data


def _detect_and_read(file_path):
    """Reads raw bytes and returns decoded lines, using chardet when available."""
    with open(file_path, "rb") as f:
        raw = f.read()

    try:
        import chardet
        detected = chardet.detect(raw)
        encoding = detected["encoding"]
        confidence = detected["confidence"]
        print(f"    Encoding: {encoding} (confidence: {confidence:.1%})")

        if confidence < 0.70:
            print("    Low confidence - switching to manual detection")
            return _fallback_decode(raw)

        return raw.decode(encoding, errors="replace").splitlines()

    except ImportError:
        print("    chardet not available - using fallback detection")
        return _fallback_decode(raw)


def _fallback_decode(raw):
    """Tries BOM markers first, then common encodings in order."""
    # BOM markers - most reliable
    if raw.startswith(b"\xff\xfe"):
        print("    Detected: UTF-16 LE (BOM)")
        return raw.decode("utf-16-le", errors="replace").splitlines()
    if raw.startswith(b"\xfe\xff"):
        print("    Detected: UTF-16 BE (BOM)")
        return raw.decode("utf-16-be", errors="replace").splitlines()
    if raw.startswith(b"\xef\xbb\xbf"):
        print("    Detected: UTF-8 (BOM)")
        return raw.decode("utf-8-sig", errors="replace").splitlines()

    for encoding in ("utf-8", "cp1252", "gbk", "big5", "shift_jis", "euc-kr"):
        try:
            text = raw.decode(encoding)
            garbage_ratio = text.count("") / max(len(text), 1)
            if garbage_ratio < 0.10:
                print(f"    Decoded with: {encoding}")
                return text.splitlines()
        except (UnicodeDecodeError, LookupError):
            continue

    print("    Using latin-1 as last resort")
    return raw.decode("latin-1", errors="replace").splitlines()


def _parse_log_lines(lines):
    """Extracts tabular data from log lines (Object header -> [END])."""
    data = []
    headers = []
    table_started = False

    for line in lines:
        clean = line.strip()

        if clean.startswith("Object"):
            headers = clean.split()
            table_started = True
            print(f"    Found headers: {len(headers)} columns")
            continue

        if clean == "[END]":
            break

        if table_started and clean:
            values = clean.split()
            if len(values) == len(headers):
                data.append(dict(zip(headers, values)))

    return data


# ============================================================================
# DATA FILTERING
# ============================================================================

def filter_data(all_data, filter_excel_path, filter_sheet_name):
    """
    Splits all_data into (kept, deleted) based on '5G cell ID' from Sheet2.
    Returns (all_data, []) if anything goes wrong.
    """
    if not all_data:
        return [], []

    try:
        filter_df = pd.read_excel(
            filter_excel_path, sheet_name=filter_sheet_name, engine="openpyxl"
        )

        if "5G cell ID" not in filter_df.columns:
            print("  Warning: '5G cell ID' column not found - keeping all data")
            return all_data, []

        cell_list = filter_df["5G cell ID"].dropna().astype(str).tolist()
        print(f"  Filter criteria: {len(cell_list)} cell IDs")

        df = pd.DataFrame(all_data)
        if "Object" not in df.columns:
            print("  Warning: 'Object' column not found in log data - keeping all data")
            return all_data, []

        kept_df    = df[df["Object"].isin(cell_list)]
        deleted_df = df[~df["Object"].isin(cell_list)]

        return kept_df.to_dict("records"), deleted_df.to_dict("records")

    except FileNotFoundError:
        print(f"  ERROR: Filter file '{filter_excel_path}' not found - keeping all data")
        return all_data, []
    except Exception as exc:
        print(f"  ERROR during filtering: {exc}")
        return all_data, []


# ============================================================================
# EXCEL OUTPUT
# ============================================================================

def output_file(data, output_excel, sheet_name):
    """Exports data (list of dicts) to the specified sheet with formatting."""
    if not data:
        print(f"  No data to write to '{sheet_name}'")
        return

    df = pd.DataFrame(data)

    # Put SourceFile first if present
    cols = df.columns.tolist()
    if "SourceFile" in cols:
        cols.insert(0, cols.pop(cols.index("SourceFile")))
    df = df[cols]

    header_fill = _make_fill(STYLE["header_fill"])
    header_font = _make_font(bold=False)
    border      = _make_border()

    try:
        writer_kwargs = dict(engine="openpyxl", mode="a", if_sheet_exists="replace")
        try:
            with pd.ExcelWriter(output_excel, **writer_kwargs) as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                _format_sheet(writer.book[sheet_name], df, header_fill, header_font, border)
        except FileNotFoundError:
            with pd.ExcelWriter(output_excel, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                _format_sheet(writer.book[sheet_name], df, header_fill, header_font, border)

        print(f"  -> Saved to '{sheet_name}': {len(data)} rows")

    except Exception as exc:
        print(f"  ERROR saving to Excel: {exc}")


def _format_sheet(ws, df, header_fill, header_font, border):
    """Applies borders, header fill, and auto column widths to a worksheet."""
    max_row = len(df) + 1
    max_col = len(df.columns)

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            if r == 1:
                cell.fill = header_fill
                cell.font = header_font

    for col_idx, col_name in enumerate(df.columns, 1):
        max_len = len(str(col_name))
        for val in df[col_name]:
            try:
                max_len = max(max_len, len(str(val)))
            except (TypeError, ValueError):
                pass
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2


# ============================================================================
# KPI VALIDATION
# ============================================================================

def check_and_format_kpi_data(file_path, sheet_name):
    """
    Reads Sheet1, marks cells that fail KPI thresholds in dark red.
    Uses KPI_RULES from config - one consistent rule set for all checks.
    """
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb[sheet_name]
    except FileNotFoundError:
        print(f"  ERROR: File '{file_path}' not found")
        return
    except KeyError:
        print(f"  ERROR: Sheet '{sheet_name}' not found")
        return

    error_fill = _make_fill(STYLE["error_fill"])
    error_font = _make_font(STYLE["error_font"])

    header = [cell.value for cell in ws[1]]

    # Build column-index lookup for each KPI rule
    rule_cols = []
    for rule in KPI_RULES:
        col_name = rule[0]
        try:
            rule_cols.append((rule, header.index(col_name) + 1))
        except ValueError:
            print(f"  WARNING: Column '{col_name}' not found in '{sheet_name}' - skipping")

    error_count = 0
    for row_idx in range(2, ws.max_row + 1):
        for rule, col_idx in rule_cols:
            cell = ws.cell(row=row_idx, column=col_idx)
            try:
                if cell.value is None:
                    continue
                value = float(cell.value)
                if not _kpi_passes(rule, value):
                    cell.fill = error_fill
                    cell.font = error_font
                    error_count += 1
            except (TypeError, ValueError):
                # Non-numeric value -> treat as fail, consistent across all KPIs
                cell.fill = error_fill
                cell.font = error_font
                error_count += 1

    wb.save(file_path)
    print(f"  -> Marked {error_count} KPI errors")


# ============================================================================
# FORM GENERATION (Sheet4)
# ============================================================================

def process_kpi_excel(filename, engineering_categories):
    """
    Generates the formatted KPI summary form in Sheet4.

    Args:
        filename (str): Path to the Excel file.
        engineering_categories (dict): {site_id: category_string}
            Collected in main.py before calling this function.
    """
    try:
        wb = load_workbook(filename)
        df_sheet1 = pd.read_excel(filename, sheet_name="Sheet1")
        df_sheet2 = pd.read_excel(filename, sheet_name="Sheet2")
    except FileNotFoundError:
        print(f"  ERROR: File '{filename}' not found")
        return
    except Exception as exc:
        print(f"  ERROR reading Excel: {exc}")
        return

    # (Re-)create Sheet4
    if "Sheet4" in wb.sheetnames:
        del wb["Sheet4"]
    ws = wb.create_sheet("Sheet4")

    thin_border = _make_border()

    col_headers = [
        {"text": "5G Site ID",                "bg": STYLE["form_header_fill"], "fg": "000000"},
        {"text": "Site Name",                  "bg": STYLE["form_header_fill"], "fg": "000000"},
        {"text": "工程類別",                   "bg": STYLE["form_header_fill"], "fg": "000000"},
    ] + [
        {"text": rule[1],                      "bg": STYLE["form_header_fill"], "fg": STYLE["form_crit_font"]}
        for rule in KPI_RULES
    ]

    # Threshold hint strings shown in the criteria row
    def _threshold_hint(rule):
        _, _, op, threshold = rule
        return f"{op} {threshold}"

    unique_site_ids = df_sheet2["5G site ID"].unique()
    print(f"  Processing {len(unique_site_ids)} sites...")

    row_offset = 1
    for idx, site_id in enumerate(unique_site_ids, 1):
        print(f"    [{idx}/{len(unique_site_ids)}] {site_id}")

        site_name_series = df_sheet2[df_sheet2["5G site ID"] == site_id]["Site Name"]
        site_name = site_name_series.iloc[0] if not site_name_series.empty else "Unknown"

        cell_ids = df_sheet2[df_sheet2["5G site ID"] == site_id]["5G cell ID"].tolist()
        df_site = df_sheet1[df_sheet1["Object"].isin(cell_ids)]

        # Evaluate KPI results for this site
        kpi_results = {}
        for rule in KPI_RULES:
            col_name = rule[0]
            failed = []
            for _, row in df_site.iterrows():
                try:
                    value = float(row[col_name])
                    if value != value:  # NaN check
                        failed.append(f"{int(row['Object'])} = N/A")
                    elif not _kpi_passes(rule, value):
                        failed.append(f"{int(row['Object'])} = {value}")
                except (ValueError, TypeError):
                    pass
            kpi_results[col_name] = "\n".join(failed) if failed else "OK"

        engineering_category = engineering_categories.get(str(site_id), "Unknown")

        # - Column headers row ------------------------
        for c, hdr in enumerate(col_headers, 1):
            cell = ws.cell(row=row_offset, column=c, value=hdr["text"])
            cell.fill      = _make_fill(hdr["bg"])
            cell.font      = Font(color=hdr["fg"], size=12, bold=True)
            cell.border    = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
        row_offset += 1

        # - Criteria hints row ------------------------
        criteria_values = (
            ["新站/mRRU/共牌/原有拆除"]           # engineering category placeholder
            + [_threshold_hint(r) for r in KPI_RULES]
        )
        # Cols 1 & 2 (site ID / site name) are merged across both data rows
        _write_cell(ws, row_offset, 1, site_id,   STYLE["form_pass_fill"], "000000", thin_border, bold=True)
        _write_cell(ws, row_offset, 2, site_name, STYLE["form_pass_fill"], "000000", thin_border, bold=True)
        for c, text in enumerate(criteria_values, 3):
            _write_cell(ws, row_offset, c, text, "ffffff", STYLE["form_note_font"], thin_border, bold=True)
        row_offset += 1

        # - Actual data row -------------------------
        _write_cell(ws, row_offset, 3, engineering_category, STYLE["form_pass_fill"], "000000", thin_border, bold=True)
        for c, rule in enumerate(KPI_RULES, 4):
            result  = kpi_results[rule[0]]
            is_pass = result == "OK"
            _write_cell(
                ws, row_offset, c, result,
                STYLE["form_pass_fill"] if is_pass else STYLE["error_fill"],
                "000000"                if is_pass else "FFFFFF",
                thin_border, bold=True,
            )
        row_offset += 1

        # Merge site ID and site name across the two data rows
        ws.merge_cells(start_row=row_offset - 2, start_column=1,
                       end_row=row_offset - 1,   end_column=1)
        ws.merge_cells(start_row=row_offset - 2, start_column=2,
                       end_row=row_offset - 1,   end_column=2)

        row_offset += 1  # blank gap between sites

    # Auto-fit column widths
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except (TypeError, ValueError):
                pass
        ws.column_dimensions[col_letter].width = max_len + 2

    wb.save(filename)
    print("  -> Sheet4 created successfully")


def _write_cell(ws, row, col, value, bg, fg, border, bold=False, size=12):
    """Helper: write a value to a cell with fill, font, border, and alignment."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill      = _make_fill(bg)
    cell.font      = Font(color=fg, size=size, bold=bold)
    cell.border    = border
    cell.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)


# ============================================================================
# REPORT GENERATION
# ============================================================================

def process_excel_template(input_excel_path):
    """
    Creates a final date-stamped report (e.g. 0115_KPI.xlsx) from the template.
    Works entirely on the in-memory workbook - saves only once.

    Returns:
        str: New filename, or None on failure.
    """
    try:
        wb_template = load_workbook(input_excel_path)

        try:
            source_sheet1 = wb_template["Sheet1"]
            source_sheet2 = wb_template["Sheet2"]
        except KeyError as exc:
            print(f"  ERROR: Sheet {exc} not found")
            return None

        # Find date in Sheet2 for the filename
        date_value = None
        for row in source_sheet2.iter_rows():
            for cell in row:
                if isinstance(cell.value, (datetime.datetime, pd.Timestamp)):
                    date_value = cell.value
                    break
            if date_value:
                break

        if not date_value:
            print("  ERROR: No valid date found in Sheet2")
            return None

        new_filename = f"{date_value.strftime('%m%d')}_KPI.xlsx"
        print(f"  -> Creating report: {new_filename}")

        # Build the new workbook in memory - no intermediate save
        new_wb = load_workbook(input_excel_path)

        for sheet_name in ("Sheet1", "Sheet2", "Sheet3"):
            if sheet_name in new_wb.sheetnames:
                del new_wb[sheet_name]
                print(f"  -> Deleted sheet: {sheet_name}")

        if "Sheet4" in new_wb.sheetnames:
            new_wb["Sheet4"].title = "Sheet"
            print("  -> Renamed 'Sheet4' to 'Sheet'")

        combined_sheet = new_wb.create_sheet(title="Sheet2")

        header_row1 = list(source_sheet1.iter_rows(min_row=1, max_row=1))[0]
        header_row2 = list(source_sheet2.iter_rows(min_row=1, max_row=1))[0]

        try:
            obj_col_idx     = next(c.column for c in header_row1 if c.value == "Object")
            cell_id_col_idx = next(c.column for c in header_row2 if c.value == "5G cell ID")

            kpi_col_indices = {}
            for rule in KPI_RULES:
                col_name, display_name = rule[0], rule[1]
                kpi_col_indices[display_name] = next(
                    c.column for c in header_row1 if c.value == col_name
                )
        except StopIteration:
            print("  ERROR: Missing required header column")
            return None

        # Final combined sheet headers
        final_headers = [
            "開台日期", "5G site ID", "Site Name", "5G cell ID",
        ] + [rule[1] for rule in KPI_RULES]

        sheet2_num_cols = len(list(source_sheet2.iter_rows(min_row=1, max_row=1))[0])

        # Write header row (copy formatting from source)
        for i, text in enumerate(final_headers, 1):
            cell = combined_sheet.cell(row=1, column=i, value=text)
            if i <= 4:
                src = source_sheet2.cell(row=1, column=i)
            else:
                display_name = final_headers[i - 1]
                src = source_sheet1.cell(row=1, column=kpi_col_indices[display_name])
            _copy_style(src, cell)

        # Build lookup: object_id -> sheet1 row
        sheet1_lookup = {
            str(row[obj_col_idx - 1].value): row
            for row in source_sheet1.iter_rows(min_row=2)
            if row[obj_col_idx - 1].value
        }

        # Write data rows
        for out_row, src_row in enumerate(source_sheet2.iter_rows(min_row=2), start=2):
            cell_id_value = str(src_row[cell_id_col_idx - 1].value)

            for i, src_cell in enumerate(src_row, 1):
                dest = combined_sheet.cell(row=out_row, column=i, value=src_cell.value)
                _copy_style(src_cell, dest)

            matching = sheet1_lookup.get(cell_id_value)
            if matching:
                for offset, (display_name, col_idx) in enumerate(kpi_col_indices.items()):
                    src_cell = matching[col_idx - 1]
                    dest = combined_sheet.cell(
                        row=out_row,
                        column=sheet2_num_cols + 1 + offset,
                        value=src_cell.value,
                    )
                    _copy_style(src_cell, dest)

        # Copy column widths
        for col_dim in source_sheet2.column_dimensions:
            combined_sheet.column_dimensions[col_dim].width = (
                source_sheet2.column_dimensions[col_dim].width
            )
        for offset, (display_name, col_idx) in enumerate(kpi_col_indices.items()):
            dest_col  = sheet2_num_cols + 1 + offset
            dest_letter = combined_sheet.cell(row=1, column=dest_col).column_letter
            src_letter  = source_sheet1.cell(row=1, column=col_idx).column_letter
            combined_sheet.column_dimensions[dest_letter].width = (
                source_sheet1.column_dimensions[src_letter].width
            )

        new_wb.save(new_filename)
        print(f"  -> Report created successfully: {new_filename}")
        return new_filename

    except FileNotFoundError:
        print(f"  ERROR: File '{input_excel_path}' not found")
        return None
    except Exception as exc:
        import traceback
        print(f"  ERROR: {exc}")
        traceback.print_exc()
        return None


def _copy_style(src_cell, dest_cell):
    """Copies font, fill, border, number format, and alignment between cells."""
    if src_cell.has_style:
        dest_cell.font         = copy(src_cell.font)
        dest_cell.fill         = copy(src_cell.fill)
        dest_cell.border       = copy(src_cell.border)
        dest_cell.number_format = src_cell.number_format
        dest_cell.alignment    = copy(src_cell.alignment)


# ============================================================================
# TEMPLATE RESET
# ============================================================================

def reset_excel_template(file_name, sheet_name_kept, sheet_name_deleted, sheet_name_form):
    """
    Clears data from kept sheet and removes deleted/form sheets.
    Prepares the template for the next run.
    """
    try:
        wb = load_workbook(file_name)

        if sheet_name_kept in wb.sheetnames:
            ws = wb[sheet_name_kept]
            for row in ws.iter_rows():
                for cell in row:
                    cell.value = None
                    cell.fill  = PatternFill(fill_type=None)
            print(f"  -> Cleaned '{sheet_name_kept}'")
        else:
            print(f"  Warning: Sheet '{sheet_name_kept}' not found")

        for name in (sheet_name_deleted, sheet_name_form):
            if name in wb.sheetnames:
                wb.remove(wb[name])
                print(f"  -> Deleted '{name}'")
            else:
                print(f"  Warning: Sheet '{name}' not found")

        wb.save(file_name)
        print("  -> Template reset complete")

    except FileNotFoundError:
        print(f"  ERROR: File '{file_name}' not found")
    except Exception as exc:
        print(f"  ERROR: {exc}")


# ============================================================================
# FILE OPENER (no review pause - removed open_and_wait)
# ============================================================================

def open_file(file_path):
    """Opens the file with the system's default application (non-blocking)."""
    try:
        if os.name == "nt":
            os.startfile(file_path)
        elif os.name == "posix":
            subprocess.Popen(["open", file_path])
        else:
            print("  Unsupported OS for auto-opening files")
    except Exception as exc:
        print(f"  ERROR opening file: {exc}")
